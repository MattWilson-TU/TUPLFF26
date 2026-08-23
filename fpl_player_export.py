#!/usr/bin/env python3
"""
Fetch FPL bootstrap-static + element-summary data and export to Excel.

Usage:
  pip install requests openpyxl
  python fpl_player_export.py
  python fpl_player_export.py --output my_players.xlsx
"""

from __future__ import annotations

import argparse
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

BOOTSTRAP_URL = "https://fantasy.premierleague.com/api/bootstrap-static/"
ELEMENT_SUMMARY_URL = "https://fantasy.premierleague.com/api/element-summary/{player_id}/"
HEADERS = {
    "User-Agent": "FPL-Player-Export/1.0 (local script; personal use)",
    "Accept": "application/json",
}


def fetch_json(session: requests.Session, url: str, retries: int = 3) -> dict:
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, headers=HEADERS, timeout=30)
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(0.5 * attempt)
    raise RuntimeError(f"Failed to fetch {url}: {last_error}") from last_error


def last_season_total(history_past: list[dict]) -> int | None:
    """Return total_points from the most recent completed season."""
    if not history_past:
        return None
    return history_past[-1].get("total_points")


def fetch_last_season_points(player_id: int) -> tuple[int, int | None]:
    # One-off request per call so concurrent workers stay thread-safe.
    with requests.Session() as session:
        data = fetch_json(session, ELEMENT_SUMMARY_URL.format(player_id=player_id))
    return player_id, last_season_total(data.get("history_past") or [])


def build_rows(bootstrap: dict, last_season_by_id: dict[int, int | None]) -> list[dict]:
    teams = {t["id"]: t["name"] for t in bootstrap["teams"]}
    positions = {t["id"]: t["singular_name"] for t in bootstrap["element_types"]}

    rows: list[dict] = []
    for player in bootstrap["elements"]:
        player_id = player["id"]
        rows.append(
            {
                "Player ID": player_id,
                "First Name": player.get("first_name", ""),
                "Surname": player.get("second_name", ""),
                "Team": teams.get(player["team"], ""),
                "Position": positions.get(player["element_type"], ""),
                "NowCost": player["now_cost"] / 10.0,
                "Last Seasons Total": last_season_by_id.get(player_id),
            }
        )
    rows.sort(key=lambda r: (r["Team"], r["Surname"], r["First Name"]))
    return rows


def write_xlsx(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "FPL Players"

    headers = [
        "Player ID",
        "First Name",
        "Surname",
        "Team",
        "Position",
        "NowCost",
        "Last Seasons Total",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row[h] for h in headers])

    # NowCost as one decimal place (e.g. 6.0 for £6.0m)
    for excel_row in range(2, len(rows) + 2):
        ws.cell(row=excel_row, column=6).number_format = "0.0"

    widths = {
        "A": 12,
        "B": 16,
        "C": 22,
        "D": 16,
        "E": 12,
        "F": 10,
        "G": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export FPL players (bootstrap + last season totals) to .xlsx"
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: fpl_players_YYYYMMDD_HHMMSS.xlsx)",
    )
    parser.add_argument(
        "-w",
        "--workers",
        type=int,
        default=12,
        help="Concurrent requests for element-summary (default: 12)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_path = args.output or Path(f"fpl_players_{stamp}.xlsx")

    print("Fetching bootstrap-static...")
    with requests.Session() as session:
        bootstrap = fetch_json(session, BOOTSTRAP_URL)
        players = bootstrap["elements"]
        player_ids = [p["id"] for p in players]
        print(f"Found {len(player_ids)} players. Fetching element-summary...")

        last_season_by_id: dict[int, int | None] = {}
        done = 0
        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {
                executor.submit(fetch_last_season_points, pid): pid
                for pid in player_ids
            }
            for future in as_completed(futures):
                player_id, points = future.result()
                last_season_by_id[player_id] = points
                done += 1
                if done % 50 == 0 or done == len(player_ids):
                    print(f"  {done}/{len(player_ids)}")

    rows = build_rows(bootstrap, last_season_by_id)
    write_xlsx(rows, output_path)
    print(f"Wrote {len(rows)} players to {output_path.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        raise SystemExit(1)
