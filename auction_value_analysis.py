#!/usr/bin/env python3
"""
FPL season-long auction value analysis.

Fetches the official FPL API, scores every player for a private auction league
(unique ownership, 11-player squads, no transfers), and writes a multi-sheet
Excel workbook with auction values, bid guidance, and bargain targets.

Usage:
  pip install -r requirements-fpl.txt
  python auction_value_analysis.py
  python auction_value_analysis.py --output my_auction_values.xlsx
  python auction_value_analysis.py --budget 150 --managers 17
  python auction_value_analysis.py --historical-file last_season.csv
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Configurable constants
# ---------------------------------------------------------------------------

# Historical season used as the primary performance baseline
TARGET_HISTORICAL_SEASON = "2025/26"

# Auction budget model (per-manager budget × managers = total room money)
# Budget is in £m (e.g. 150 = £150m per manager)
TOTAL_LEAGUE_BUDGET = 150
NUMBER_OF_MANAGERS = 17
SQUAD_SIZE = 11

# Fraction of total room budget assumed to be spent on the drafted pool
BUDGET_SPEND_RATIO = 0.95

# Power used when converting auction_value → share of budget (>1 concentrates
# spend on the best players; 1.0 is strictly proportional)
PRICE_ALLOCATION_POWER = 1.35

# Bid band multipliers around suggested_auction_price
MIN_BID_FACTOR = 0.75
TARGET_BID_FACTOR = 1.00
MAX_BID_FACTOR = 1.30

# Nominal price floor / replacement-level scaling for players outside the
# expected drafted pool (top managers × squad size)
OUTSIDE_POOL_PRICE_FACTOR = 0.45
ABSOLUTE_MIN_BID = 1.0

# Auction-value component weights (must sum to 1.0)
WEIGHT_POINTS_PERCENTILE = 0.30
WEIGHT_PRICE_EFFICIENCY = 0.20
WEIGHT_RELIABILITY = 0.15
WEIGHT_EXPECTED_POINTS = 0.15
WEIGHT_POINTS_PER_90 = 0.10
WEIGHT_OVERLOOKED = 0.10

# How strongly position scarcity reshapes the final 0–100 auction_value
SCARCITY_BLEND = 0.12

# Reliability / confidence thresholds
FULL_SEASON_MINUTES = 3420  # 38 × 90
HIGH_CONFIDENCE_MINUTES = 1500
MEDIUM_CONFIDENCE_MINUTES = 500
ROTATION_FRIENDLY_POSITIONS = {"Goalkeeper", "Midfielder"}

# API
BOOTSTRAP_URL = "https://fantasy.premierleague.com/api/bootstrap-static/"
ELEMENT_SUMMARY_URL = (
    "https://fantasy.premierleague.com/api/element-summary/{player_id}/"
)
HEADERS = {
    "User-Agent": "FPL-Auction-Value-Analysis/1.0 (local script; personal use)",
    "Accept": "application/json",
}
DEFAULT_OUTPUT = Path(
    "/Users/matt/Desktop/FPL26-27 data/fpl_auction_values_2026_27.xlsx"
)
DEFAULT_WORKERS = 12
REQUEST_RETRIES = 3

POSITION_ORDER = ["Goalkeeper", "Defender", "Midfielder", "Forward"]
POSITION_SHORT = {
    "Goalkeeper": "GKP",
    "Defender": "DEF",
    "Midfielder": "MID",
    "Forward": "FWD",
}

logger = logging.getLogger("auction_value_analysis")


# ---------------------------------------------------------------------------
# API / I/O helpers
# ---------------------------------------------------------------------------


def fetch_json(session: requests.Session, url: str, retries: int = REQUEST_RETRIES) -> dict:
    """GET JSON with basic retries. Raises RuntimeError on persistent failure."""
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(url, headers=HEADERS, timeout=30)
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(0.5 * attempt)
    raise RuntimeError(f"Failed to fetch {url}: {last_error}") from last_error


def fetch_bootstrap_data(session: requests.Session | None = None) -> dict:
    """Download bootstrap-static (players, teams, positions, events)."""
    owns_session = session is None
    session = session or requests.Session()
    try:
        logger.info("Fetching bootstrap-static…")
        data = fetch_json(session, BOOTSTRAP_URL)
        logger.info(
            "Bootstrap loaded: %s players, %s teams",
            len(data.get("elements", [])),
            len(data.get("teams", [])),
        )
        return data
    finally:
        if owns_session:
            session.close()


def _history_for_season(history_past: list[dict], season: str) -> dict | None:
    """Return the history_past row for a named season, if present."""
    for row in history_past or []:
        if row.get("season_name") == season:
            return row
    return None


def fetch_player_history(
    player_id: int, season: str = TARGET_HISTORICAL_SEASON
) -> tuple[int, dict[str, Any]]:
    """
    Fetch element-summary for one player and extract the target season.

    Returns (player_id, stats_dict). On failure or missing season, stats
    fields are None rather than raising.
    """
    empty: dict[str, Any] = {
        "last_season_points": None,
        "minutes": None,
        "appearances": None,
        "history_error": None,
        "season_found": False,
    }
    try:
        with requests.Session() as session:
            data = fetch_json(session, ELEMENT_SUMMARY_URL.format(player_id=player_id))
        past = data.get("history_past") or []
        row = _history_for_season(past, season)
        if row is None:
            # Fall back to most recent completed season if exact name missing
            if past:
                row = past[-1]
                if row.get("season_name") != season:
                    empty["history_error"] = (
                        f"season {season} not found; nearest={row.get('season_name')}"
                    )
                    return player_id, empty
            else:
                empty["history_error"] = "no history_past"
                return player_id, empty

        starts = row.get("starts")
        minutes = row.get("minutes")
        # Appearances: prefer starts; if absent, estimate from minutes
        appearances = starts
        if appearances is None and minutes is not None and minutes > 0:
            appearances = max(1, int(round(minutes / 90.0)))

        return player_id, {
            "last_season_points": row.get("total_points"),
            "minutes": minutes,
            "appearances": appearances,
            "starts": starts,
            "season_name": row.get("season_name"),
            "season_found": True,
            "history_error": None,
            "raw_history": row,
        }
    except Exception as exc:  # noqa: BLE001 — per-player isolation
        empty["history_error"] = str(exc)
        return player_id, empty


def fetch_historical_data(
    player_ids: list[int],
    workers: int = DEFAULT_WORKERS,
    season: str = TARGET_HISTORICAL_SEASON,
) -> dict[int, dict[str, Any]]:
    """Concurrently fetch previous-season stats for all player IDs."""
    logger.info(
        "Fetching element-summary for %s players (%s workers, season=%s)…",
        len(player_ids),
        workers,
        season,
    )
    results: dict[int, dict[str, Any]] = {}
    done = 0
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(fetch_player_history, pid, season): pid for pid in player_ids
        }
        for future in as_completed(futures):
            pid = futures[future]
            try:
                player_id, stats = future.result()
                results[player_id] = stats
            except Exception as exc:  # noqa: BLE001
                logger.warning("History fetch failed for %s: %s", pid, exc)
                results[pid] = {
                    "last_season_points": None,
                    "minutes": None,
                    "appearances": None,
                    "history_error": str(exc),
                    "season_found": False,
                }
            done += 1
            if done % 50 == 0 or done == len(player_ids):
                logger.info("  history %s/%s", done, len(player_ids))
    return results


def load_historical_file(path: Path) -> dict[int, dict[str, Any]]:
    """
    Load historical stats from a local CSV or JSON file.

    CSV columns (minimum): player_id, total_points
    Optional: minutes, appearances, starts

    JSON: list of objects with the same keys, or a dict keyed by player_id.
    """
    logger.info("Loading historical data from %s", path)
    if not path.exists():
        raise FileNotFoundError(f"Historical file not found: {path}")

    records: list[dict[str, Any]]
    if path.suffix.lower() == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            records = [
                {"player_id": int(k), **(v if isinstance(v, dict) else {"total_points": v})}
                for k, v in payload.items()
            ]
        elif isinstance(payload, list):
            records = payload
        else:
            raise ValueError("JSON historical file must be a list or object")
    else:
        df = pd.read_csv(path)
        records = df.to_dict(orient="records")

    out: dict[int, dict[str, Any]] = {}
    for row in records:
        try:
            pid = int(row["player_id"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"Invalid historical row (need player_id): {row}") from exc
        points = row.get("total_points", row.get("last_season_points"))
        minutes = row.get("minutes")
        appearances = row.get("appearances", row.get("starts"))
        out[pid] = {
            "last_season_points": None if points is None or (isinstance(points, float) and math.isnan(points)) else int(points),
            "minutes": None if minutes is None or (isinstance(minutes, float) and math.isnan(minutes)) else int(minutes),
            "appearances": None
            if appearances is None or (isinstance(appearances, float) and math.isnan(appearances))
            else int(appearances),
            "season_found": points is not None,
            "history_error": None,
            "season_name": row.get("season_name", TARGET_HISTORICAL_SEASON),
        }
    logger.info("Loaded historical stats for %s players from file", len(out))
    return out


# ---------------------------------------------------------------------------
# Dataset construction
# ---------------------------------------------------------------------------


def _safe_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _na_display(value: Any) -> Any:
    """Excel-friendly missing marker."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "N/A"
    try:
        if pd.isna(value):
            return "N/A"
    except (TypeError, ValueError):
        pass
    return value


def _round_series(series: pd.Series, decimals: int = 1) -> pd.Series:
    """Round a series that may contain pandas NA / None without raising."""
    numeric = pd.to_numeric(series, errors="coerce")
    return numeric.round(decimals)


def build_player_dataset(
    bootstrap: dict,
    historical: dict[int, dict[str, Any]],
) -> pd.DataFrame:
    """Merge bootstrap current-season fields with previous-season history."""
    teams = {t["id"]: t for t in bootstrap.get("teams", [])}
    positions = {
        t["id"]: t.get("singular_name", str(t["id"]))
        for t in bootstrap.get("element_types", [])
    }

    rows: list[dict[str, Any]] = []
    for player in bootstrap.get("elements", []):
        pid = int(player["id"])
        team = teams.get(player.get("team"), {})
        hist = historical.get(pid, {})

        # Team strength proxy (API often zeros early in a season)
        soh = _safe_float(team.get("strength_overall_home"), 0.0) or 0.0
        soa = _safe_float(team.get("strength_overall_away"), 0.0) or 0.0
        strength = (soh + soa) / 2.0 if (soh or soa) else _safe_float(team.get("strength"), 0.0) or 0.0

        rows.append(
            {
                "player_id": pid,
                "player": f"{player.get('first_name', '')} {player.get('second_name', '')}".strip()
                or player.get("web_name", ""),
                "web_name": player.get("web_name", ""),
                "team": team.get("name", ""),
                "team_short": team.get("short_name", ""),
                "position": positions.get(player.get("element_type"), "Unknown"),
                "element_type": player.get("element_type"),
                "current_fpl_price": (player.get("now_cost") or 0) / 10.0,
                "status": player.get("status", ""),
                "chance_of_playing_next_round": player.get("chance_of_playing_next_round"),
                "news": player.get("news") or "",
                "selected_by_percent": _safe_float(player.get("selected_by_percent"), 0.0) or 0.0,
                "current_season_points": player.get("total_points"),
                "current_season_minutes": player.get("minutes"),
                "form": _safe_float(player.get("form"), 0.0),
                "ep_next": _safe_float(player.get("ep_next"), 0.0),
                "team_strength": strength,
                "last_season_points": hist.get("last_season_points"),
                "minutes": hist.get("minutes"),
                "appearances": hist.get("appearances"),
                "season_found": bool(hist.get("season_found")),
                "history_error": hist.get("history_error"),
                "season_name": hist.get("season_name"),
            }
        )

    df = pd.DataFrame(rows)
    logger.info("Built dataset for %s players", len(df))
    return df


# ---------------------------------------------------------------------------
# Metric calculations
# ---------------------------------------------------------------------------


def _percentile_rank(series: pd.Series) -> pd.Series:
    """Percentile rank 0–100; NaNs stay NaN."""
    if series.notna().sum() == 0:
        return pd.Series([float("nan")] * len(series), index=series.index)
    return series.rank(pct=True, method="average") * 100.0


def _clip_0_100(series: pd.Series) -> pd.Series:
    return series.clip(lower=0, upper=100)


def calculate_performance_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Historical output rates: pts/app, pts/90, pts per current £m."""
    out = df.copy()
    pts = pd.to_numeric(out["last_season_points"], errors="coerce")
    apps = pd.to_numeric(out["appearances"], errors="coerce")
    mins = pd.to_numeric(out["minutes"], errors="coerce")
    price = pd.to_numeric(out["current_fpl_price"], errors="coerce")

    out["points_per_appearance"] = pts / apps.replace(0, pd.NA)
    out["points_per_90"] = pts / (mins / 90.0).replace(0, pd.NA)
    # Guard absurd pts/90 from tiny minute samples
    low_mins = mins.isna() | (mins < 180)
    out.loc[low_mins, "points_per_90"] = pd.NA

    out["points_per_current_fpl_price"] = pts / price.replace(0, pd.NA)
    return out


def calculate_reliability(df: pd.DataFrame) -> pd.DataFrame:
    """
    Reliability 0–100 from minutes, appearances, and availability status.

    Rotation-heavy positions (GKP/MID) are penalised less for sub-max minutes.
    """
    out = df.copy()
    scores: list[float] = []

    for _, row in out.iterrows():
        mins = row["minutes"]
        apps = row["appearances"]
        status = (row.get("status") or "").lower()
        position = row.get("position") or ""

        if mins is None or (isinstance(mins, float) and math.isnan(mins)):
            # No history → low reliability, further cut if currently unavailable
            base = 25.0
            if status in {"i", "s", "u", "n"}:
                base = 10.0
            scores.append(base)
            continue

        mins_f = float(mins)
        apps_f = float(apps) if apps is not None and not (isinstance(apps, float) and math.isnan(apps)) else mins_f / 90.0

        minutes_component = min(100.0, (mins_f / FULL_SEASON_MINUTES) * 100.0)
        apps_component = min(100.0, (apps_f / 38.0) * 100.0)

        # Completeness when selected: average minutes per appearance
        if apps_f > 0:
            mpg = mins_f / apps_f
            completeness = min(100.0, (mpg / 90.0) * 100.0)
        else:
            completeness = 0.0

        # Soften the minutes bar for positions that routinely rotate
        if position in ROTATION_FRIENDLY_POSITIONS:
            minutes_component = min(100.0, minutes_component * 1.15)

        score = (
            0.45 * minutes_component
            + 0.35 * apps_component
            + 0.20 * completeness
        )

        # Current unavailability soft penalty (season-long auction still cares)
        if status == "i":
            score *= 0.85
        elif status in {"s", "u"}:
            score *= 0.70
        elif status == "n":
            score *= 0.55
        elif status == "d":
            chance = row.get("chance_of_playing_next_round")
            if chance is not None:
                score *= 0.85 + 0.15 * (float(chance) / 100.0)

        # Sporadic appearances with decent headline minutes are rare; catch
        # the opposite: very few apps
        if apps_f < 10 and mins_f < 900:
            score *= 0.75

        scores.append(float(_clip_0_100(pd.Series([score])).iloc[0]))

    out["reliability_score"] = scores
    return out


def calculate_consistency(df: pd.DataFrame) -> pd.DataFrame:
    """
    Consistency 0–100 from aggregate last-season patterns.

    GW-by-GW variance is not available once a new season starts, so this uses
    starts/minutes regularity as an objective proxy.
    """
    out = df.copy()
    values: list[float] = []
    for _, row in out.iterrows():
        mins = row["minutes"]
        apps = row["appearances"]
        if mins is None or apps is None:
            values.append(float("nan"))
            continue
        try:
            mins_f = float(mins)
            apps_f = float(apps)
        except (TypeError, ValueError):
            values.append(float("nan"))
            continue
        if apps_f <= 0 or mins_f <= 0:
            values.append(0.0)
            continue

        participation = min(1.0, apps_f / 38.0)
        mpg = mins_f / apps_f
        role_stability = min(1.0, mpg / 80.0)  # ~80+ mins/app ⇒ settled role
        volume = min(1.0, mins_f / 2500.0)
        score = 100.0 * (0.40 * participation + 0.35 * role_stability + 0.25 * volume)
        values.append(max(0.0, min(100.0, score)))

    out["consistency_score"] = values
    return out


def calculate_position_scarcity(df: pd.DataFrame) -> pd.DataFrame:
    """
    Derive scarcity from the actual points distribution in each position.

    Position scarcity is high when top-end points are rare relative to the
    median (steep drop-off). Player scarcity blends that with the player's
    own standing inside the position.
    """
    out = df.copy()
    out["scarcity_score"] = float("nan")
    out["position_scarcity"] = float("nan")

    position_scores: dict[str, float] = {}

    for position, group in out.groupby("position"):
        pts = pd.to_numeric(group["last_season_points"], errors="coerce").dropna()
        # Require a meaningful sample of players with real minutes
        mins = pd.to_numeric(group.loc[pts.index, "minutes"], errors="coerce")
        pts = pts[mins.fillna(0) >= 180]
        if len(pts) < 5:
            position_scores[position] = 50.0
            continue

        p50 = float(pts.quantile(0.50))
        p75 = float(pts.quantile(0.75))
        p90 = float(pts.quantile(0.90))
        top = float(pts.max())

        # Steepness: how far the elite sit above a replaceable starter
        drop_90_50 = (p90 / p50) if p50 > 0 else 1.0
        drop_top_75 = (top / p75) if p75 > 0 else 1.0
        # Rarity: share of players within 15% of the position's 90th percentile
        near_elite = float((pts >= p90 * 0.85).mean()) if p90 > 0 else 1.0
        rarity = 1.0 - near_elite  # higher when few players are near elite

        # Combine into a raw score; normalised across positions below
        raw = (0.45 * drop_90_50) + (0.25 * drop_top_75) + (0.30 * (1.0 + rarity))
        position_scores[position] = raw

    raw_vals = list(position_scores.values())
    lo, hi = (min(raw_vals), max(raw_vals)) if raw_vals else (0.0, 1.0)

    def norm_pos(raw: float) -> float:
        if hi <= lo:
            return 50.0
        return 20.0 + 60.0 * ((raw - lo) / (hi - lo))  # map into ~20–80

    normalised = {pos: norm_pos(raw) for pos, raw in position_scores.items()}
    logger.info("Position scarcity (derived): %s", {k: round(v, 1) for k, v in normalised.items()})

    scarcity_col: list[float] = []
    pos_scar_col: list[float] = []
    for _, row in out.iterrows():
        pos = row["position"]
        pos_scar = normalised.get(pos, 50.0)
        pos_scar_col.append(pos_scar)

        pts = row["last_season_points"]
        if pts is None or (isinstance(pts, float) and math.isnan(pts)):
            scarcity_col.append(pos_scar * 0.4)
            continue

        # Player standing within position
        peers = pd.to_numeric(
            out.loc[out["position"] == pos, "last_season_points"], errors="coerce"
        )
        if peers.notna().sum() == 0:
            scarcity_col.append(pos_scar * 0.5)
            continue
        standing = float((peers <= float(pts)).mean() * 100.0)  # percentile 0–100
        # High scarcity only matters if the player is actually good
        player_scar = 0.55 * pos_scar + 0.45 * standing
        scarcity_col.append(max(0.0, min(100.0, player_scar)))

    out["position_scarcity"] = pos_scar_col
    out["scarcity_score"] = scarcity_col
    return out


def calculate_position_comparisons(df: pd.DataFrame) -> pd.DataFrame:
    """Points vs position average/median and price-implied value gap."""
    out = df.copy()
    out["points_percentile_in_position"] = float("nan")
    out["points_vs_position_average"] = float("nan")
    out["points_vs_position_median"] = float("nan")
    out["value_vs_current_price"] = float("nan")
    out["price_efficiency_percentile"] = float("nan")

    for position, idx in out.groupby("position").groups.items():
        subset = out.loc[idx]
        pts = pd.to_numeric(subset["last_season_points"], errors="coerce")
        price = pd.to_numeric(subset["current_fpl_price"], errors="coerce")
        ppp = pd.to_numeric(subset["points_per_current_fpl_price"], errors="coerce")

        avg = pts.mean(skipna=True)
        med = pts.median(skipna=True)
        out.loc[idx, "points_percentile_in_position"] = _percentile_rank(pts)
        out.loc[idx, "points_vs_position_average"] = pts - avg
        out.loc[idx, "points_vs_position_median"] = pts - med
        out.loc[idx, "price_efficiency_percentile"] = _percentile_rank(ppp)

        # Implied points from a simple within-position price regression:
        # E[pts] ≈ slope * price, anchored through the origin of replaceable
        # production (median pts / median price).
        valid = pts.notna() & price.notna() & (price > 0)
        if valid.sum() >= 5:
            med_price = float(price[valid].median())
            med_pts = float(pts[valid].median())
            implied_per_million = (med_pts / med_price) if med_price else 0.0
            implied = price * implied_per_million
            out.loc[idx, "value_vs_current_price"] = pts - implied
        else:
            out.loc[idx, "value_vs_current_price"] = pts - pts.mean(skipna=True)

    return out


def calculate_overlooked_value(df: pd.DataFrame) -> pd.DataFrame:
    """
    Overlooked-value 0–100: strong history + efficiency + reliability at a
    price / ownership profile that suggests the market may under-rate them.
    """
    out = df.copy()
    out["overlooked_value_score"] = 0.0

    for _, idx in out.groupby("position").groups.items():
        subset = out.loc[idx]
        pts_pct = pd.to_numeric(subset["points_percentile_in_position"], errors="coerce")
        eff_pct = pd.to_numeric(subset["price_efficiency_percentile"], errors="coerce")
        rel = pd.to_numeric(subset["reliability_score"], errors="coerce")
        price = pd.to_numeric(subset["current_fpl_price"], errors="coerce")
        own = pd.to_numeric(subset["selected_by_percent"], errors="coerce")
        pts = pd.to_numeric(subset["last_season_points"], errors="coerce")

        # Low price relative to position (cheaper ⇒ more overlooked potential)
        price_pct = _percentile_rank(price)  # high = expensive
        cheapness = 100.0 - price_pct.fillna(50.0)

        # Low ownership = lower public profile (objective proxy)
        own_pct = _percentile_rank(own.fillna(0.0))
        low_profile = 100.0 - own_pct.fillna(50.0)

        # Elite-points bonus: in top tier of position points but not priced as such
        top_tier = (pts_pct.fillna(0.0) >= 80.0).astype(float) * 100.0

        for i in subset.index:
            if pd.isna(pts.loc[i]):
                score = 0.0
            else:
                score = (
                    0.30 * float(pts_pct.loc[i] if pd.notna(pts_pct.loc[i]) else 0.0)
                    + 0.25 * float(eff_pct.loc[i] if pd.notna(eff_pct.loc[i]) else 0.0)
                    + 0.15 * float(rel.loc[i] if pd.notna(rel.loc[i]) else 0.0)
                    + 0.15 * float(cheapness.loc[i] if pd.notna(cheapness.loc[i]) else 50.0)
                    + 0.10 * float(low_profile.loc[i] if pd.notna(low_profile.loc[i]) else 50.0)
                    + 0.05 * float(top_tier.loc[i])
                )
            out.loc[i, "overlooked_value_score"] = max(0.0, min(100.0, score))

    return out


def calculate_expected_points(df: pd.DataFrame) -> pd.DataFrame:
    """
    Expected-points score 0–100 within each position.

    Foundation = last-season points percentile, with modest adjustments for
    reliability, pts/90, price efficiency, team strength, and early-season
    importance signals (form / ep_next) when present.
    """
    out = df.copy()
    out["expected_points_score"] = float("nan")

    # Team strength normalised 0–1 across the league
    strength = pd.to_numeric(out["team_strength"], errors="coerce").fillna(0.0)
    if strength.max() > strength.min():
        strength_norm = (strength - strength.min()) / (strength.max() - strength.min())
    else:
        strength_norm = pd.Series(0.5, index=out.index)

    for position, idx in out.groupby("position").groups.items():
        subset = out.loc[idx]
        base = pd.to_numeric(subset["points_percentile_in_position"], errors="coerce").fillna(0.0)
        rel = pd.to_numeric(subset["reliability_score"], errors="coerce").fillna(40.0) / 100.0
        p90 = pd.to_numeric(subset["points_per_90"], errors="coerce")
        p90_pct = _percentile_rank(p90).fillna(50.0) / 100.0
        eff = pd.to_numeric(subset["price_efficiency_percentile"], errors="coerce").fillna(50.0) / 100.0
        team_c = strength_norm.loc[idx]

        form = pd.to_numeric(subset["form"], errors="coerce").fillna(0.0)
        ep = pd.to_numeric(subset["ep_next"], errors="coerce").fillna(0.0)
        # Only nudge if the API is actually producing signal
        form_pct = (_percentile_rank(form).fillna(50.0) / 100.0) if form.sum() > 0 else pd.Series(0.5, index=idx)
        ep_pct = (_percentile_rank(ep).fillna(50.0) / 100.0) if ep.sum() > 0 else pd.Series(0.5, index=idx)
        importance = 0.5 * form_pct + 0.5 * ep_pct

        # Weights keep history dominant (~70% base)
        adjusted = (
            0.70 * base
            + 0.10 * (rel * 100.0)
            + 0.08 * (p90_pct * 100.0)
            + 0.07 * (eff * 100.0)
            + 0.03 * (team_c * 100.0)
            + 0.02 * (importance * 100.0)
        )
        # Players with no history stay low-confidence low scores
        no_hist = pd.to_numeric(subset["last_season_points"], errors="coerce").isna()
        adjusted = adjusted.where(~no_hist, adjusted.clip(upper=25.0))
        out.loc[idx, "expected_points_score"] = _clip_0_100(adjusted)

    return out


def assign_confidence(df: pd.DataFrame) -> pd.DataFrame:
    """HIGH / MEDIUM / LOW confidence from volume of historical evidence."""
    out = df.copy()
    labels: list[str] = []
    for _, row in out.iterrows():
        if not row.get("season_found") or row["last_season_points"] is None:
            labels.append("LOW")
            continue
        mins = row["minutes"]
        if mins is None or (isinstance(mins, float) and math.isnan(mins)):
            labels.append("LOW")
            continue
        mins_f = float(mins)
        if mins_f >= HIGH_CONFIDENCE_MINUTES and float(row.get("appearances") or 0) >= 15:
            labels.append("HIGH")
        elif mins_f >= MEDIUM_CONFIDENCE_MINUTES:
            labels.append("MEDIUM")
        else:
            labels.append("LOW")
    out["confidence"] = labels
    return out


def calculate_auction_value(df: pd.DataFrame) -> pd.DataFrame:
    """
    Final auction_value 0–100, computed separately within each position.

    Components are percentile-normalised inside the position before weighting.
    Scarcity applies a light post-blend so rare high-end output is rewarded
    without drowning historical performance.
    """
    weights = {
        "points": WEIGHT_POINTS_PERCENTILE,
        "efficiency": WEIGHT_PRICE_EFFICIENCY,
        "reliability": WEIGHT_RELIABILITY,
        "expected": WEIGHT_EXPECTED_POINTS,
        "p90": WEIGHT_POINTS_PER_90,
        "overlooked": WEIGHT_OVERLOOKED,
    }
    weight_sum = sum(weights.values())
    if abs(weight_sum - 1.0) > 1e-6:
        logger.warning("Auction-value weights sum to %s (expected 1.0)", weight_sum)

    out = df.copy()
    out["auction_value"] = float("nan")
    # Expose normalised components for auditability
    for key in weights:
        out[f"component_{key}"] = float("nan")

    for position, idx in out.groupby("position").groups.items():
        subset = out.loc[idx]

        c_points = pd.to_numeric(subset["points_percentile_in_position"], errors="coerce").fillna(0.0)
        c_eff = pd.to_numeric(subset["price_efficiency_percentile"], errors="coerce").fillna(0.0)
        c_rel = _percentile_rank(pd.to_numeric(subset["reliability_score"], errors="coerce")).fillna(0.0)
        c_exp = _percentile_rank(pd.to_numeric(subset["expected_points_score"], errors="coerce")).fillna(0.0)
        c_p90 = _percentile_rank(pd.to_numeric(subset["points_per_90"], errors="coerce")).fillna(0.0)
        c_ov = _percentile_rank(pd.to_numeric(subset["overlooked_value_score"], errors="coerce")).fillna(0.0)

        base = (
            weights["points"] * c_points
            + weights["efficiency"] * c_eff
            + weights["reliability"] * c_rel
            + weights["expected"] * c_exp
            + weights["p90"] * c_p90
            + weights["overlooked"] * c_ov
        )

        scarcity = pd.to_numeric(subset["scarcity_score"], errors="coerce").fillna(50.0)
        # Blend: auction_value = (1-s)*base + s*base*(0.85 + 0.30*scarcity/100)
        scarcity_factor = 0.85 + 0.30 * (scarcity / 100.0)
        blended = (1.0 - SCARCITY_BLEND) * base + SCARCITY_BLEND * (base * scarcity_factor)
        blended = _clip_0_100(blended)

        # Suppress scores for no-history players so they don't pollute ranks
        no_hist = pd.to_numeric(subset["last_season_points"], errors="coerce").isna()
        blended = blended.where(~no_hist, blended * 0.35)

        out.loc[idx, "auction_value"] = blended
        out.loc[idx, "component_points"] = c_points
        out.loc[idx, "component_efficiency"] = c_eff
        out.loc[idx, "component_reliability"] = c_rel
        out.loc[idx, "component_expected"] = c_exp
        out.loc[idx, "component_p90"] = c_p90
        out.loc[idx, "component_overlooked"] = c_ov

    return out


def calculate_auction_prices(
    df: pd.DataFrame,
    budget_per_manager: float = TOTAL_LEAGUE_BUDGET,
    managers: int = NUMBER_OF_MANAGERS,
    squad_size: int = SQUAD_SIZE,
) -> pd.DataFrame:
    """
    Distribute total room budget across players using auction_value shares.

    Transparent model:
      total_budget = budget_per_manager × managers
      roster_spots = managers × squad_size
      Among the top `roster_spots` players by auction_value:
          weight_i = auction_value_i ** PRICE_ALLOCATION_POWER
          suggested_i = total_budget × BUDGET_SPEND_RATIO × weight_i / Σ weights
      Players outside that pool get a scaled replacement-level price so the
      sheet still has guidance without inventing demand that will not exist.
    """
    out = df.copy()
    total_budget = budget_per_manager * managers
    roster_spots = managers * squad_size
    spend_pool = total_budget * BUDGET_SPEND_RATIO

    values = pd.to_numeric(out["auction_value"], errors="coerce").fillna(0.0)
    # Rank globally for draft-pool membership (managers buy across positions)
    ranked_idx = values.sort_values(ascending=False).index.tolist()
    draft_idx = ranked_idx[:roster_spots]
    outside_idx = ranked_idx[roster_spots:]

    suggested = pd.Series(0.0, index=out.index)

    if draft_idx:
        draft_vals = values.loc[draft_idx].clip(lower=0.01)
        weights = draft_vals ** PRICE_ALLOCATION_POWER
        shares = weights / weights.sum()
        suggested.loc[draft_idx] = shares * spend_pool
        marginal_price = float(suggested.loc[draft_idx].min())
        marginal_value = float(draft_vals.min())
    else:
        marginal_price = ABSOLUTE_MIN_BID
        marginal_value = 1.0

    for i in outside_idx:
        v = float(values.loc[i])
        if v <= 0:
            suggested.loc[i] = ABSOLUTE_MIN_BID
        else:
            ratio = v / marginal_value if marginal_value else 0.0
            suggested.loc[i] = max(
                ABSOLUTE_MIN_BID,
                marginal_price * ratio * OUTSIDE_POOL_PRICE_FACTOR,
            )

    out["suggested_auction_price"] = suggested.round(1)
    out["minimum_recommended_bid"] = (suggested * MIN_BID_FACTOR).clip(lower=ABSOLUTE_MIN_BID).round(1)
    out["target_bid"] = (suggested * TARGET_BID_FACTOR).clip(lower=ABSOLUTE_MIN_BID).round(1)
    out["maximum_bid"] = (suggested * MAX_BID_FACTOR).clip(lower=ABSOLUTE_MIN_BID).round(1)

    out.attrs["total_budget"] = total_budget
    out.attrs["roster_spots"] = roster_spots
    out.attrs["spend_pool"] = spend_pool
    return out


def classify_players(df: pd.DataFrame) -> pd.DataFrame:
    """Derive value_rating and auction_strategy from metrics (no name rules)."""
    out = df.copy()
    ratings: list[str] = []
    strategies: list[str] = []

    for _, row in out.iterrows():
        av = float(row["auction_value"]) if pd.notna(row["auction_value"]) else 0.0
        ov = float(row["overlooked_value_score"]) if pd.notna(row["overlooked_value_score"]) else 0.0
        eff = float(row["price_efficiency_percentile"]) if pd.notna(row.get("price_efficiency_percentile")) else 0.0
        conf = row.get("confidence", "LOW")
        pts = row["last_season_points"]
        status = (row.get("status") or "").lower()

        if pts is None or (isinstance(pts, float) and math.isnan(pts)) or conf == "LOW" and av < 40:
            if status in {"u", "n", "s"}:
                rating, strategy = "AVOID", "AVOID"
            elif av >= 50:
                rating, strategy = "FAIR", "WATCH"
            else:
                rating, strategy = "AVOID", "AVOID"
        elif av >= 90 and ov >= 70:
            rating, strategy = "ELITE", "TARGET"
        elif av >= 85:
            rating, strategy = "ELITE", "TARGET"
        elif av >= 75 and (ov >= 65 or eff >= 70):
            rating, strategy = "EXCELLENT VALUE", "TARGET"
        elif av >= 70:
            rating, strategy = "EXCELLENT VALUE", "TARGET IF CHEAP"
        elif av >= 55 and ov >= 60:
            rating, strategy = "GOOD VALUE", "TARGET IF CHEAP"
        elif av >= 55:
            rating, strategy = "GOOD VALUE", "WATCH"
        elif av >= 40:
            rating, strategy = "FAIR", "DO NOT OVERPAY"
        elif av >= 25:
            rating, strategy = "OVERPRICED", "DO NOT OVERPAY"
        else:
            rating, strategy = "AVOID", "AVOID"

        # Injured / unavailable: never TARGET
        if status in {"i", "d"} and strategy == "TARGET":
            strategy = "TARGET IF CHEAP"
        if status in {"u", "n", "s"}:
            strategy = "AVOID"
            if rating not in {"AVOID", "OVERPRICED"}:
                rating = "OVERPRICED"

        ratings.append(rating)
        strategies.append(strategy)

    out["value_rating"] = ratings
    out["auction_strategy"] = strategies
    return out


def calculate_all_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Run the full metric pipeline in dependency order."""
    out = calculate_performance_metrics(df)
    out = calculate_reliability(out)
    out = calculate_consistency(out)
    out = calculate_position_scarcity(out)
    out = calculate_position_comparisons(out)
    out = calculate_overlooked_value(out)
    out = calculate_expected_points(out)
    out = assign_confidence(out)
    out = calculate_auction_value(out)
    out = calculate_auction_prices(out)
    out = classify_players(out)
    return out


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def _style_header(ws, row: int = 1) -> None:
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 32
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 28) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells[:80]:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_width, max(10, length + 2))


def _excel_value(value: Any) -> Any:
    """Convert pandas/numpy missing values into Excel-safe None / N/A."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return "N/A"
    except (TypeError, ValueError):
        pass
    # numpy types
    if hasattr(value, "item") and not isinstance(value, (bytes, str)):
        try:
            return value.item()
        except (ValueError, AttributeError):
            pass
    return value


def _write_dataframe(ws, df: pd.DataFrame) -> None:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_excel_value(value) if r_idx > 1 else value)
    if df.shape[0] >= 0 and df.shape[1] > 0:
        _style_header(ws)
        _autosize(ws)


def _apply_value_rating_formatting(ws, rating_col: int, start_row: int, end_row: int) -> None:
    fills = {
        "ELITE": PatternFill("solid", fgColor="1B5E20"),
        "EXCELLENT VALUE": PatternFill("solid", fgColor="2E7D32"),
        "GOOD VALUE": PatternFill("solid", fgColor="81C784"),
        "FAIR": PatternFill("solid", fgColor="FFF59D"),
        "OVERPRICED": PatternFill("solid", fgColor="FFB74D"),
        "AVOID": PatternFill("solid", fgColor="E57373"),
    }
    letter = get_column_letter(rating_col)
    for label, fill in fills.items():
        font = Font(
            bold=True,
            color="FFFFFF" if label in {"ELITE", "EXCELLENT VALUE"} else "000000",
        )
        ws.conditional_formatting.add(
            f"{letter}{start_row}:{letter}{end_row}",
            CellIsRule(operator="equal", formula=[f'"{label}"'], fill=fill, font=font),
        )


def _auction_board_frame(df: pd.DataFrame) -> pd.DataFrame:
    board = df.copy()
    # Sort by position order then auction value
    board["_pos_order"] = board["position"].map(
        {p: i for i, p in enumerate(POSITION_ORDER)}
    ).fillna(99)
    board = board.sort_values(
        ["_pos_order", "auction_value"], ascending=[True, False]
    ).reset_index(drop=True)
    board.insert(0, "Rank", board.groupby("position").cumcount() + 1)

    display = pd.DataFrame(
        {
            "Rank": board["Rank"],
            "Position": board["position"].map(lambda p: POSITION_SHORT.get(p, p)),
            "Player": board["player"],
            "Team": board["team"],
            "Last Season Points": board["last_season_points"].map(_na_display),
            "Appearances": board["appearances"].map(_na_display),
            "Minutes": board["minutes"].map(_na_display),
            "Points / Appearance": _round_series(board["points_per_appearance"], 2).map(_na_display),
            "Points / 90": _round_series(board["points_per_90"], 2).map(_na_display),
            "Current FPL Price": _round_series(board["current_fpl_price"], 1),
            "Points / £m": _round_series(board["points_per_current_fpl_price"], 2).map(_na_display),
            "Reliability Score": _round_series(board["reliability_score"], 1),
            "Expected Points Score": _round_series(board["expected_points_score"], 1).map(_na_display),
            "Overlooked Value Score": _round_series(board["overlooked_value_score"], 1),
            "Auction Value": _round_series(board["auction_value"], 1),
            "Suggested Auction Price": board["suggested_auction_price"],
            "Minimum Bid": board["minimum_recommended_bid"],
            "Target Bid": board["target_bid"],
            "Maximum Bid": board["maximum_bid"],
            "Value Rating": board["value_rating"],
            "Auction Strategy": board["auction_strategy"],
            "Points vs Pos Avg": _round_series(board["points_vs_position_average"], 1).map(_na_display),
            "Points vs Pos Median": _round_series(board["points_vs_position_median"], 1).map(_na_display),
            "Value vs Current Price": _round_series(board["value_vs_current_price"], 1).map(_na_display),
            "Consistency Score": _round_series(board["consistency_score"], 1).map(_na_display),
            "Scarcity Score": _round_series(board["scarcity_score"], 1),
            "Confidence": board["confidence"],
            "Pts Percentile (Pos)": _round_series(board["points_percentile_in_position"], 1).map(_na_display),
            "Price Eff Percentile": _round_series(board["price_efficiency_percentile"], 1).map(_na_display),
            "Comp: Points": _round_series(board["component_points"], 1),
            "Comp: Efficiency": _round_series(board["component_efficiency"], 1),
            "Comp: Reliability": _round_series(board["component_reliability"], 1),
            "Comp: Expected": _round_series(board["component_expected"], 1),
            "Comp: Pts/90": _round_series(board["component_p90"], 1),
            "Comp: Overlooked": _round_series(board["component_overlooked"], 1),
            "Ownership %": _round_series(board["selected_by_percent"], 1),
            "Status": board["status"],
        }
    )
    return display


def create_excel_workbook(
    df: pd.DataFrame,
    output_path: Path,
    budget_per_manager: float,
    managers: int,
    squad_size: int,
) -> None:
    """Write all five worksheets to output_path."""
    wb = Workbook()

    # --- Sheet 1: Auction Board ---
    ws1 = wb.active
    ws1.title = "Auction Board"
    board = _auction_board_frame(df)
    _write_dataframe(ws1, board)
    rating_col = list(board.columns).index("Value Rating") + 1
    if len(board) > 0:
        _apply_value_rating_formatting(ws1, rating_col, 2, len(board) + 1)
        # Highlight high auction values
        av_col = list(board.columns).index("Auction Value") + 1
        av_letter = get_column_letter(av_col)
        ws1.conditional_formatting.add(
            f"{av_letter}2:{av_letter}{len(board)+1}",
            FormulaRule(
                formula=[f"AND({av_letter}2>=80,{av_letter}2<>\"N/A\")"],
                fill=PatternFill("solid", fgColor="C8E6C9"),
            ),
        )

    # --- Sheet 2: Position Rankings (top 20 each) ---
    ws2 = wb.create_sheet("Position Rankings")
    ws2["A1"] = "Top 20 players by Auction Value in each position"
    ws2["A1"].font = Font(bold=True, size=14)
    start_row = 3
    for position in POSITION_ORDER:
        subset = (
            df[df["position"] == position]
            .sort_values("auction_value", ascending=False)
            .head(20)
            .copy()
        )
        subset.insert(0, "Rank", range(1, len(subset) + 1))
        block = pd.DataFrame(
            {
                "Rank": subset["Rank"],
                "Player": subset["player"],
                "Team": subset["team"],
                "Last Season Points": subset["last_season_points"].map(_na_display),
                "Current FPL Price": _round_series(subset["current_fpl_price"], 1),
                "Auction Value": _round_series(subset["auction_value"], 1),
                "Overlooked Value": _round_series(subset["overlooked_value_score"], 1),
                "Target Bid": subset["target_bid"],
                "Value Rating": subset["value_rating"],
                "Strategy": subset["auction_strategy"],
                "Confidence": subset["confidence"],
            }
        )
        ws2.cell(row=start_row, column=1, value=position).font = Font(bold=True, size=12)
        start_row += 1
        for r_idx, row in enumerate(dataframe_to_rows(block, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = ws2.cell(
                    row=r_idx,
                    column=c_idx,
                    value=value if r_idx == start_row else _excel_value(value),
                )
                if r_idx == start_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9E2F3")
        start_row += len(block) + 3
    _autosize(ws2)

    # --- Sheet 3: Bargains ---
    ws3 = wb.create_sheet("Bargains")
    bargains = (
        df.sort_values("overlooked_value_score", ascending=False)
        .head(30)
        .copy()
    )
    bargains.insert(0, "Rank", range(1, len(bargains) + 1))
    bargain_df = pd.DataFrame(
        {
            "Rank": bargains["Rank"],
            "Position": bargains["position"].map(lambda p: POSITION_SHORT.get(p, p)),
            "Player": bargains["player"],
            "Team": bargains["team"],
            "Last Season Points": bargains["last_season_points"].map(_na_display),
            "Current FPL Price": _round_series(bargains["current_fpl_price"], 1),
            "Points / £m": _round_series(bargains["points_per_current_fpl_price"], 2).map(_na_display),
            "Ownership %": _round_series(bargains["selected_by_percent"], 1),
            "Reliability": _round_series(bargains["reliability_score"], 1),
            "Overlooked Value Score": _round_series(bargains["overlooked_value_score"], 1),
            "Auction Value": _round_series(bargains["auction_value"], 1),
            "Target Bid": bargains["target_bid"],
            "Maximum Bid": bargains["maximum_bid"],
            "Value Rating": bargains["value_rating"],
            "Strategy": bargains["auction_strategy"],
            "Value vs Current Price": _round_series(bargains["value_vs_current_price"], 1).map(_na_display),
            "Confidence": bargains["confidence"],
        }
    )
    _write_dataframe(ws3, bargain_df)
    if len(bargain_df) > 0:
        ov_col = list(bargain_df.columns).index("Overlooked Value Score") + 1
        _apply_value_rating_formatting(
            ws3,
            list(bargain_df.columns).index("Value Rating") + 1,
            2,
            len(bargain_df) + 1,
        )
        letter = get_column_letter(ov_col)
        ws3.conditional_formatting.add(
            f"{letter}2:{letter}{len(bargain_df)+1}",
            FormulaRule(
                formula=[f"{letter}2>=75"],
                fill=PatternFill("solid", fgColor="A5D6A7"),
            ),
        )

    # --- Sheet 4: Methodology ---
    ws4 = wb.create_sheet("Methodology")
    total_budget = budget_per_manager * managers
    roster_spots = managers * squad_size
    methodology_lines = [
        "FPL Auction Value Analysis — Methodology",
        "",
        "PURPOSE",
        "Estimate how much to bid for each player in a season-long private auction",
        "where each manager buys 11 unique players and holds them all season.",
        "This is NOT a reproduction of official FPL prices.",
        "",
        "DATA SOURCES",
        f"Bootstrap: {BOOTSTRAP_URL}",
        f"Per-player history: {ELEMENT_SUMMARY_URL}",
        f"Primary historical season: {TARGET_HISTORICAL_SEASON} (last_season_points)",
        "Optional override: --historical-file CSV/JSON with player_id, total_points,",
        "minutes, appearances/starts.",
        "Missing history is shown as N/A; confidence is lowered.",
        "",
        "CONFIGURABLE AUCTION CONSTANTS",
        f"TOTAL_LEAGUE_BUDGET (per manager) = {budget_per_manager}",
        f"NUMBER_OF_MANAGERS = {managers}",
        f"SQUAD_SIZE = {squad_size}",
        f"Total room budget = {total_budget}",
        f"Expected roster spots filled = {roster_spots}",
        f"BUDGET_SPEND_RATIO = {BUDGET_SPEND_RATIO}",
        f"PRICE_ALLOCATION_POWER = {PRICE_ALLOCATION_POWER}",
        f"MIN_BID_FACTOR = {MIN_BID_FACTOR}",
        f"TARGET_BID_FACTOR = {TARGET_BID_FACTOR}",
        f"MAX_BID_FACTOR = {MAX_BID_FACTOR}",
        f"OUTSIDE_POOL_PRICE_FACTOR = {OUTSIDE_POOL_PRICE_FACTOR}",
        f"SCARCITY_BLEND = {SCARCITY_BLEND}",
        "",
        "AUCTION VALUE WEIGHTS (within each position)",
        f"{WEIGHT_POINTS_PERCENTILE:.0%} previous-season points percentile",
        f"{WEIGHT_PRICE_EFFICIENCY:.0%} points / current FPL price efficiency",
        f"{WEIGHT_RELIABILITY:.0%} reliability",
        f"{WEIGHT_EXPECTED_POINTS:.0%} expected points score",
        f"{WEIGHT_POINTS_PER_90:.0%} points per 90",
        f"{WEIGHT_OVERLOOKED:.0%} overlooked-value score",
        "Each component is percentile-normalised within the player's position",
        "before weighting. Scores are NOT compared across positions.",
        "",
        "INDIVIDUAL METRICS",
        "last_season_points — total FPL points in the target historical season",
        "appearances — starts from history_past (best available appearance proxy)",
        "minutes — total minutes in that season",
        "points_per_appearance — points / appearances",
        "points_per_90 — points / (minutes/90); suppressed if minutes < 180",
        "points_per_current_fpl_price — last-season points / current FPL £m price",
        "reliability_score — 0–100 from minutes, appearances, completeness, status",
        "consistency_score — regularity proxy from participation and mins/appearance",
        "expected_points_score — history-led score with modest adjustments for",
        "  reliability, pts/90, price efficiency, team strength, form/ep_next",
        "overlooked_value_score — high history + efficiency + reliability at low",
        "  price / ownership (public-profile proxies; no social scraping)",
        "points_vs_position_average / median — points gap vs same-position peers",
        "value_vs_current_price — points minus price-implied points",
        "  (implied = price × position median pts / median price)",
        "scarcity_score — derived from how steeply points fall off in the position",
        "  and how high the player sits in that distribution",
        "confidence — HIGH / MEDIUM / LOW from historical minute volume",
        "",
        "SUGGESTED AUCTION PRICE MODEL",
        "1. Take the top (managers × squad_size) players by auction_value.",
        "2. weight_i = auction_value_i ** PRICE_ALLOCATION_POWER",
        "3. suggested_i = total_budget × BUDGET_SPEND_RATIO × weight_i / Σ weights",
        "4. Outside that pool: scaled replacement-level price",
        "   (marginal_draft_price × value_ratio × OUTSIDE_POOL_PRICE_FACTOR).",
        "5. minimum = suggested × MIN_BID_FACTOR",
        "   target   = suggested × TARGET_BID_FACTOR",
        "   maximum  = suggested × MAX_BID_FACTOR",
        "This forces realistic spending across 11-player squads rather than",
        "dumping the entire budget on a single elite name.",
        "",
        "VALUE RATING & STRATEGY",
        "Generated from auction_value, overlooked_value, efficiency, confidence,",
        "and current availability status — never from hard-coded player names.",
        "Ratings: ELITE, EXCELLENT VALUE, GOOD VALUE, FAIR, OVERPRICED, AVOID",
        "Strategies: TARGET, TARGET IF CHEAP, WATCH, DO NOT OVERPAY, AVOID",
        "",
        "ROBUSTNESS",
        "Per-player API failures are isolated. New / promoted / low-minute players",
        "receive LOW confidence and suppressed auction values instead of invented",
        "history. All component columns are included on Auction Board for audit.",
    ]
    for i, line in enumerate(methodology_lines, start=1):
        ws4.cell(row=i, column=1, value=line)
        if i == 1:
            ws4.cell(row=i, column=1).font = Font(bold=True, size=14)
        elif line in {
            "PURPOSE",
            "DATA SOURCES",
            "CONFIGURABLE AUCTION CONSTANTS",
            "AUCTION VALUE WEIGHTS (within each position)",
            "INDIVIDUAL METRICS",
            "SUGGESTED AUCTION PRICE MODEL",
            "VALUE RATING & STRATEGY",
            "ROBUSTNESS",
        }:
            ws4.cell(row=i, column=1).font = Font(bold=True, color="1F4E79")
    ws4.column_dimensions["A"].width = 100

    # --- Sheet 5: Raw Data ---
    ws5 = wb.create_sheet("Raw Data")
    raw_cols = [
        "player_id",
        "player",
        "web_name",
        "team",
        "team_short",
        "position",
        "current_fpl_price",
        "status",
        "chance_of_playing_next_round",
        "news",
        "selected_by_percent",
        "current_season_points",
        "current_season_minutes",
        "form",
        "ep_next",
        "team_strength",
        "season_name",
        "season_found",
        "last_season_points",
        "minutes",
        "appearances",
        "history_error",
        "points_per_appearance",
        "points_per_90",
        "points_per_current_fpl_price",
        "reliability_score",
        "consistency_score",
        "expected_points_score",
        "overlooked_value_score",
        "points_percentile_in_position",
        "price_efficiency_percentile",
        "points_vs_position_average",
        "points_vs_position_median",
        "value_vs_current_price",
        "position_scarcity",
        "scarcity_score",
        "auction_value",
        "suggested_auction_price",
        "minimum_recommended_bid",
        "target_bid",
        "maximum_bid",
        "value_rating",
        "auction_strategy",
        "confidence",
        "component_points",
        "component_efficiency",
        "component_reliability",
        "component_expected",
        "component_p90",
        "component_overlooked",
    ]
    raw = df[[c for c in raw_cols if c in df.columns]].copy()
    _write_dataframe(ws5, raw)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Wrote workbook to %s", output_path.resolve())


# ---------------------------------------------------------------------------
# CLI / main
# ---------------------------------------------------------------------------


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyse FPL players and produce auction-value Excel guidance."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--budget",
        type=float,
        default=TOTAL_LEAGUE_BUDGET,
        help=f"Budget per manager (default: {TOTAL_LEAGUE_BUDGET})",
    )
    parser.add_argument(
        "--managers",
        type=int,
        default=NUMBER_OF_MANAGERS,
        help=f"Number of managers (default: {NUMBER_OF_MANAGERS})",
    )
    parser.add_argument(
        "--squad-size",
        type=int,
        default=SQUAD_SIZE,
        help=f"Players per manager (default: {SQUAD_SIZE})",
    )
    parser.add_argument(
        "--historical-file",
        type=Path,
        default=None,
        help="Optional CSV/JSON of previous-season stats (overrides API history)",
    )
    parser.add_argument(
        "--season",
        type=str,
        default=TARGET_HISTORICAL_SEASON,
        help=f"Historical season name to use (default: {TARGET_HISTORICAL_SEASON})",
    )
    parser.add_argument(
        "-w",
        "--workers",
        type=int,
        default=DEFAULT_WORKERS,
        help=f"Concurrent element-summary requests (default: {DEFAULT_WORKERS})",
    )
    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
        help="Debug logging",
    )
    return parser.parse_args(argv)


def print_summary(df: pd.DataFrame, output_path: Path) -> None:
    """Concise end-of-run summary for the terminal."""
    print()
    print("FPL Auction Value Analysis complete.")
    print()
    print(f"Players analysed: {len(df)}")
    print()
    print("Top targets:")
    targets = (
        df[df["auction_strategy"].isin(["TARGET", "TARGET IF CHEAP"])]
        .sort_values("auction_value", ascending=False)
        .head(8)
    )
    if targets.empty:
        targets = df.sort_values("auction_value", ascending=False).head(5)
    for _, row in targets.iterrows():
        pos = POSITION_SHORT.get(row["position"], row["position"])
        print(
            f"  {pos:3}  {row['player']:<28}  {row['auction_value']:5.1f}  "
            f"Target £{row['target_bid']:.0f}"
        )
    print()
    print("Best overlooked values:")
    overlooked = df.sort_values("overlooked_value_score", ascending=False).head(5)
    for _, row in overlooked.iterrows():
        pts = row["last_season_points"]
        pts_s = "N/A" if pts is None or (isinstance(pts, float) and math.isnan(pts)) else str(int(pts))
        print(
            f"  {row['player']:<28}  {pts_s:>4} pts  "
            f"£{row['current_fpl_price']:.1f}m  "
            f"OV {row['overlooked_value_score']:.0f}  "
            f"AV {row['auction_value']:.0f}  "
            f"Target £{row['target_bid']:.0f}"
        )
    print()
    print("Output:")
    print(f"  {output_path.resolve()}")
    print()


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s %(message)s",
    )

    if args.managers < 1 or args.squad_size < 1 or args.budget <= 0:
        logger.error("budget, managers and squad-size must be positive")
        return 2

    try:
        bootstrap = fetch_bootstrap_data()
        player_ids = [int(p["id"]) for p in bootstrap.get("elements", [])]
        if not player_ids:
            logger.error("No players returned from bootstrap-static")
            return 1

        if args.historical_file:
            historical = load_historical_file(args.historical_file)
            # Fill any gaps from the API so new players still appear
            missing = [pid for pid in player_ids if pid not in historical]
            if missing:
                logger.info(
                    "Historical file missing %s players; fetching those from API…",
                    len(missing),
                )
                historical.update(
                    fetch_historical_data(missing, workers=args.workers, season=args.season)
                )
        else:
            historical = fetch_historical_data(
                player_ids, workers=args.workers, season=args.season
            )

        df = build_player_dataset(bootstrap, historical)
        df = calculate_all_metrics(df)
        # Recalculate prices with CLI budget overrides
        df = calculate_auction_prices(
            df,
            budget_per_manager=args.budget,
            managers=args.managers,
            squad_size=args.squad_size,
        )
        df = classify_players(df)

        create_excel_workbook(
            df,
            args.output,
            budget_per_manager=args.budget,
            managers=args.managers,
            squad_size=args.squad_size,
        )
        print_summary(df, args.output)
        return 0
    except KeyboardInterrupt:
        print("\nCancelled.", file=sys.stderr)
        return 130
    except Exception as exc:  # noqa: BLE001
        logger.exception("Analysis failed: %s", exc)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
